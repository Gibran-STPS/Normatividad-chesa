#!/usr/bin/env python3
"""Genera un Excel de auditoría con todas las bases consolidadas (respaldo documental)."""
import json, argparse, os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

HEADER_FILL = PatternFill('solid', start_color='1C2024', end_color='1C2024')
HEADER_FONT = Font(color='F1EFE9', bold=True, name='Arial', size=10)
BODY_FONT = Font(name='Arial', size=10)
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='1C2024')
AREA_RESPONSABLE = 'Responsable de Auditoria STPS'
MESES = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto',
         'septiembre','octubre','noviembre','diciembre']

def dataset_periodo(dataset):
    """Determina el periodo (mes/año) global a partir de las fechas reales cargadas en el
    dataset — no de un valor fijo — para que el respaldo refleje automáticamente el mes de
    los archivos que se suban cada vez, sin necesidad de tocar el código cada mes."""
    import re
    conteo = {}
    for r in dataset.get('registros_diarios', []):
        if not r.get('fecha'):
            continue
        key = str(r['fecha'])[:7]
        if re.match(r'^\d{4}-\d{2}$', key):
            conteo[key] = conteo.get(key, 0) + 1
    if not conteo:
        for f in dataset.get('faltas_resumen', []):
            for fecha in (f.get('fechas') or []):
                key = str(fecha)[:7]
                if re.match(r'^\d{4}-\d{2}$', key):
                    conteo[key] = conteo.get(key, 0) + 1
    anio_ref = dataset.get('anio_referencia', 'S/D')
    if not conteo:
        return {'label': f'Fase {anio_ref}', 'file_tag': f'SD_{anio_ref}'}
    key = max(conteo, key=conteo.get)
    anio, mes_num = key.split('-')
    idx = int(mes_num) - 1
    mes_nombre = MESES[idx] if 0 <= idx < len(MESES) else ''
    mes_cap = mes_nombre.capitalize() if mes_nombre else ''
    return {'label': f'{mes_cap} {anio}'.strip(), 'file_tag': f'{(mes_nombre or "SD").upper()}_{anio}'}


def style_sheet(ws, df, title, logo_path=None):
    ws.insert_rows(1, 2)
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns),1))
    ws['A2'] = f'Grupo Chesa · Elaborado por {AREA_RESPONSABLE}'
    ws['A2'].font = Font(name='Arial', size=8.5, italic=True, color='6B675F')
    if logo_path and os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.height = 34
        img.width = 34 * (img.width / img.height) if img.height else img.width
        col_letter = get_column_letter(max(len(df.columns), 1) + 1)
        ws.add_image(img, f'{col_letter}1')
    header_row = 3
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=header_row, column=j)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
    for j, col in enumerate(df.columns, start=1):
        maxlen = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).values[:200]])
        ws.column_dimensions[get_column_letter(j)].width = min(max(12, maxlen + 2), 45)
    ws.freeze_panes = ws.cell(row=header_row+1, column=1)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dataset', required=True)
    ap.add_argument('--kpis', required=True)
    ap.add_argument('--out', required=True)
    ap.add_argument('--logo', default=None, help='Ruta al logo de Chesa (png)')
    ap.add_argument('--elabora', default=None, help='Nombre de quien elabora (opcional, para prellenar la bitácora de firmas)')
    ap.add_argument('--recibe', default=None, help='Nombre de quien recibe (opcional)')
    ap.add_argument('--vobo', default=None, help='Nombre de quien da el visto bueno (opcional)')
    args = ap.parse_args()

    with open(args.dataset, encoding='utf-8') as f:
        dataset = json.load(f)
    with open(args.kpis, encoding='utf-8') as f:
        kpis = json.load(f)

    with pd.ExcelWriter(args.out, engine='openpyxl') as xw:
        df_u = pd.DataFrame(kpis['unidades'])
        cols_u = ['unidad_negocio','marca','sucursal','total_colaboradores','colaboradores_marcaje_completo',
                  'horas_totales','horas_promedio_colaborador','semanas_totales','semanas_cumple',
                  'pct_cumplimiento_semanal','semanas_excede_tope_extra','exceso_diario',
                  'checadas_incompletas','retardo_dias','retardo_6_12','retardo_13_24',
                  'retardo_mas_25','faltas_total']
        df_u = df_u[cols_u]
        df_u.columns = ['Unidad de Negocio','Marca','Sucursal','Colaboradores','Colaboradores Marcaje Completo',
                         'Horas Totales Trabajadas','Prom. Horas/Colaborador','Semanas-Colaborador Totales','Semanas Dentro del Límite',
                         '% Cumplimiento Semanal','Semanas que Exceden Tope Legal Extra','Jornadas >12h/día',
                         'Checadas Incompletas','Días con Retardo','Retardo 6-12min','Retardo 13-24min',
                         'Retardo +25min','Faltas Totales']
        df_u.to_excel(xw, sheet_name='Resumen Unidades', index=False, startrow=2)

        df_a = pd.DataFrame(kpis['areas'])
        df_a.columns = [c.replace('_',' ').title() for c in df_a.columns]
        df_a.to_excel(xw, sheet_name='Resumen Areas', index=False, startrow=2)

        df_ed = pd.DataFrame(kpis['incidencias']['exceso_diario'])
        if not df_ed.empty:
            df_ed.columns = [c.replace('_',' ').title() for c in df_ed.columns]
        df_ed.to_excel(xw, sheet_name='Exceso Diario +12h', index=False, startrow=2)

        df_es = pd.DataFrame(kpis['incidencias']['exceso_semanal'])
        if not df_es.empty:
            df_es.columns = [c.replace('_',' ').title() for c in df_es.columns]
        df_es.to_excel(xw, sheet_name='Exceso Semanal', index=False, startrow=2)

        df_ci = pd.DataFrame(kpis['incidencias']['checadas_incompletas'])
        if not df_ci.empty:
            df_ci['fechas'] = df_ci['fechas'].apply(lambda x: ', '.join(x))
            df_ci.columns = [c.replace('_',' ').title() for c in df_ci.columns]
        df_ci.to_excel(xw, sheet_name='Checadas Incompletas', index=False, startrow=2)

        df_f = pd.DataFrame(kpis['incidencias']['faltas'])
        if not df_f.empty:
            df_f['fechas'] = df_f['fechas'].apply(lambda x: ', '.join(x))
            df_f.columns = [c.replace('_',' ').title() for c in df_f.columns]
        df_f.to_excel(xw, sheet_name='Faltas', index=False, startrow=2)

        df_r = pd.DataFrame(kpis['incidencias']['retardos'])
        if not df_r.empty:
            df_r.columns = [c.replace('_',' ').title() for c in df_r.columns]
        df_r.to_excel(xw, sheet_name='Retardos', index=False, startrow=2)

        df_sfh = pd.DataFrame(kpis['incidencias'].get('sabados_fuera_horario', []))
        if not df_sfh.empty:
            df_sfh.columns = [c.replace('_',' ').title() for c in df_sfh.columns]
        df_sfh.to_excel(xw, sheet_name='Sabados Fuera de Horario', index=False, startrow=2)

        df_diario = pd.DataFrame(dataset['registros_diarios'])
        df_diario.columns = [c.replace('_',' ').title() for c in df_diario.columns]
        df_diario.to_excel(xw, sheet_name='Detalle Diario Completo', index=False, startrow=2)

        df_sem = pd.DataFrame(dataset['semanal'])
        df_sem.columns = [c.replace('_',' ').title() for c in df_sem.columns]
        df_sem.to_excel(xw, sheet_name='Detalle Semanal Completo', index=False, startrow=2)

        sim = kpis.get('simulador_transicion')
        if sim and sim.get('detalle_por_unidad'):
            df_sim = pd.DataFrame(sim['detalle_por_unidad'])
            df_sim.columns = ['Unidad de Negocio', 'Prom. Horas/Semana Actual',
                               f"% Cumpliría con Límite {sim['anio_meta']} ({sim['limite_meta']}h)",
                               f"Brecha (h/semana) vs {sim['anio_meta']}"]
        else:
            df_sim = pd.DataFrame(columns=['Unidad de Negocio', 'Prom. Horas/Semana Actual', '% Cumpliría', 'Brecha (h/semana)'])
        df_sim.to_excel(xw, sheet_name='Simulador Transicion', index=False, startrow=2)

        unidades_keys = [u['unidad_negocio'] for u in kpis['unidades']]
        periodo = dataset_periodo(dataset)
        df_bit = pd.DataFrame([{
            'Unidad': u, 'Reporte': f'Reporte_Cumplimiento_{u}_{periodo["file_tag"]}.pdf',
            'Elabora': args.elabora or '', 'Fecha Elabora': '' if not args.elabora else '(por firmar)',
            'Recibe': args.recibe or '', 'Fecha Recibe': '' if not args.recibe else '(por firmar)',
            'Visto Bueno': args.vobo or '', 'Fecha Visto Bueno': '' if not args.vobo else '(por firmar)',
            'Observaciones': '',
        } for u in unidades_keys])
        df_bit.to_excel(xw, sheet_name='Bitacora Firmas', index=False, startrow=2)

        wb = xw.book
        titles = {
            'Resumen Unidades': f'Resumen de Cumplimiento por Unidad de Negocio — {periodo["label"]}',
            'Resumen Areas': 'Resumen por Área / Departamento',
            'Exceso Diario +12h': 'Jornadas Mayores a 12 Horas en un Día (Art. 68 LFT)',
            'Exceso Semanal': 'Semanas por Encima del Límite Legal Vigente (Art. 59 LFT)',
            'Checadas Incompletas': 'Colaboradores con Checadas Incompletas (Omisión Entrada/Salida)',
            'Faltas': 'Faltas Registradas (Sin Checada en el Día)',
            'Retardos': 'Incumplimiento Puntual — Retardos por Colaborador',
            'Sabados Fuera de Horario': 'Sábados Fuera del Horario Oficial de Grupo (09:00–14:00, sin comida)',
            'Detalle Diario Completo': 'Base Completa de Registros Diarios (Evidencia Documental)',
            'Detalle Semanal Completo': 'Base Completa de Horas por Semana y Colaborador',
            'Simulador Transicion': f"Simulador de Transición — Brecha por Unidad hacia {kpis.get('simulador_transicion',{}).get('anio_meta','2030')}",
            'Bitacora Firmas': 'Bitácora de Firmas — Elabora / Recibe / Visto Bueno',
        }
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            df_map = {
                'Resumen Unidades': df_u, 'Resumen Areas': df_a, 'Exceso Diario +12h': df_ed,
                'Exceso Semanal': df_es, 'Checadas Incompletas': df_ci, 'Faltas': df_f,
                'Retardos': df_r, 'Detalle Diario Completo': df_diario, 'Detalle Semanal Completo': df_sem,
                'Simulador Transicion': df_sim, 'Bitacora Firmas': df_bit, 'Sabados Fuera de Horario': df_sfh,
            }
            style_sheet(ws, df_map[sheet_name], titles[sheet_name], logo_path=args.logo)

    print('Excel de auditoría escrito en', args.out)

if __name__ == '__main__':
    main()
