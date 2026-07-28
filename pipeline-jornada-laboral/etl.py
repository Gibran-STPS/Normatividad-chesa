#!/usr/bin/env python3
"""
ETL - Reloj Checador Grupo Automotriz
Procesa N archivos REPORTE_DE_RELOJ_CHECADOR_<MARCA>_<SUCURSAL>_<MES>_<ANIO>.xlsx
y genera un dataset consolidado (JSON) para alimentar el dashboard y los reportes PDF.

Uso:
    python3 etl.py --data-dir ./data --config config_normativo.json --out ./output/dataset.json
"""
import json, re, sys, argparse, glob, os
from datetime import datetime, timedelta
import openpyxl

MESES = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
         'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
NON_DATA_SHEETS = {'RESUMEN RETARDOS', 'DIAS SIN CHECADAS'}
# Palabras de relleno que pueden aparecer al inicio del nombre del archivo en cualquier
# combinación/orden ("Reporte de Reloj Checador", "Reportes_Reloj_Checador", etc.) — se
# quitan una por una en vez de exigir una frase exacta, para tolerar variaciones mes a mes.
FILLER_PREFIX_WORDS = {'REPORTE', 'REPORTES', 'DE', 'RELOJ', 'CHECADOR', 'CHECADORES', 'DEL', 'LOS', 'LAS'}
# Marcas que opera el grupo — sirven de ancla si, después de quitar el relleno, el primer
# token todavía no es una marca válida (por ejemplo si el nombre trae alguna palabra extra
# que no está en FILLER_PREFIX_WORDS).
MARCAS_CONOCIDAS = {'NISSAN', 'RENAULT', 'CHANGAN'}

def parse_filename(path):
    name = os.path.basename(path).replace('.xlsx', '')
    parts = [p for p in re.split(r'[\s_]+', name) if p != '']
    while parts and parts[0].upper() in FILLER_PREFIX_WORDS:
        parts.pop(0)
    anio = None
    mes = None
    for i in range(len(parts) - 1, -1, -1):
        if parts[i].isdigit() and len(parts[i]) == 4:
            anio = parts[i]
            if i > 0 and parts[i-1].upper() in MESES:
                mes = parts[i-1].upper()
                marca_sucursal = parts[:i-1]
            else:
                marca_sucursal = parts[:i]
            break
    if anio is None:
        marca_sucursal = parts
        anio = 'S/D'
        mes = 'S/D'
    if marca_sucursal and marca_sucursal[0].upper() not in MARCAS_CONOCIDAS:
        for i, tok in enumerate(marca_sucursal):
            if tok.upper() in MARCAS_CONOCIDAS:
                marca_sucursal = marca_sucursal[i:]
                break
    marca = marca_sucursal[0] if marca_sucursal else 'S/D'
    sucursal = '_'.join(marca_sucursal[1:]) if len(marca_sucursal) > 1 else marca
    unidad = '_'.join(marca_sucursal)
    return {
        'marca': marca,
        'sucursal': sucursal,
        'unidad_negocio': unidad,
        'mes': mes,
        'anio': anio
    }

def strip_accents(s):
    import unicodedata
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def es_sabado(dia_semana, fecha):
    """Detecta sábado por el texto de la columna DIA (tolerante a acentos/mayúsculas) y,
    si ese texto no ayuda, por la fecha real del registro."""
    if dia_semana:
        if 'SAB' in strip_accents(str(dia_semana)).strip().upper():
            return True
    try:
        return datetime.strptime(str(fecha), '%Y-%m-%d').weekday() == 5
    except Exception:
        return False

def clasifica_retardo_minutos(mins_tarde):
    """Misma escala de tolerancia que ya usa el sistema de checador (6-12 / 13-24 / +25 min)."""
    if mins_tarde is None or mins_tarde <= 5:
        return False, False, False
    if mins_tarde <= 12:
        return True, False, False
    if mins_tarde <= 24:
        return False, True, False
    return False, False, True

def es_nombre_placeholder(nombre):
    """Algunos archivos traen registros sin colaborador real identificado, con nombres
    placeholder tipo 'NN-1', 'NN 2', 'NN' (y variantes como 'SIN NOMBRE', 'S/N', 'N/A') en vez
    de un nombre real. Esos registros no representan a un colaborador y no deben contarse como
    tal en ningún KPI ni listado."""
    if not nombre:
        return True
    n = str(nombre).strip().upper()
    if not n:
        return True
    if re.match(r'^NN[\s\-_]*\d*$', n):
        return True
    if n in ('SIN NOMBRE', 'S/N', 'N/A', 'NO IDENTIFICADO', 'SIN IDENTIFICAR'):
        return True
    return False

def normalize_emp_id(v):
    """Excel a veces guarda una columna de ID 'numérica' como flotante (170024 -> 170024.0),
    y si una hoja del mismo archivo la formatea distinto que otra, el mismo colaborador queda
    con dos IDs distintos y no se puede cruzar su información (checadas, retardos, faltas).
    Aquí se homologan a una sola forma de texto."""
    s = str(v if v is not None else '').strip()
    if re.match(r'^\d+\.0+$', s):
        s = re.sub(r'\.0+$', '', s)
    return s

def to_minutes(hhmm):
    if hhmm in (None, '0', 0, ''):
        return None
    try:
        h, m = str(hhmm).split(':')
        return int(h) * 60 + int(m)
    except Exception:
        return None

def horas_trabajadas(entrada, salida, s_comer, r_comer):
    e = to_minutes(entrada)
    s = to_minutes(salida)
    if e is None or s is None:
        return None
    total = s - e
    if total < 0:
        total += 24 * 60  # turno cruza medianoche
    sc = to_minutes(s_comer)
    rc = to_minutes(r_comer)
    if sc is not None and rc is not None:
        comida = rc - sc
        if comida < 0:
            comida += 24 * 60
        if 0 < comida < total:
            total -= comida
    return round(total / 60.0, 2)

def find_header_row(ws):
    for r in range(1, 5):
        vals = [c.value for c in ws[r]]
        if vals and 'ID' in vals and 'NOMBRE' in vals and 'FECHA' in vals:
            return r, vals
    raise ValueError('No se encontró fila de encabezado ID/NOMBRE/FECHA')

def col_index(headers, name):
    try:
        return headers.index(name)
    except ValueError:
        return None

def load_area_sheet(ws, area_name, meta, config=None):
    config = config or {}
    sab_cfg = config.get('jornada_sabado', {})
    sab_activo = bool(sab_cfg.get('activo', False))
    sab_entrada = sab_cfg.get('hora_entrada', '09:00')
    sab_salida = sab_cfg.get('hora_salida', '14:00')
    sab_con_comida = bool(sab_cfg.get('con_comida', False))
    sab_tolerancia = int(sab_cfg.get('tolerancia_salida_minutos', 15))
    sab_horario_label = f'{sab_entrada} - {sab_salida}'
    sab_entrada_min = to_minutes(sab_entrada)
    sab_salida_min = to_minutes(sab_salida)

    header_row, headers = find_header_row(ws)
    idx = {h: col_index(headers, h) for h in
           ['ID','NOMBRE','FECHA','DIA','HORARIO','HORARIO COMIDA','ENTRADA',
            'S. COMER','R. COMER','SALIDA','R 6-12','R 13-24','R +25']}
    records = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[idx['ID']] is None and row[idx['NOMBRE']] is None:
            continue
        def get(col):
            i = idx.get(col)
            return row[i] if i is not None and i < len(row) else None
        emp_id = normalize_emp_id(get('ID'))
        nombre = str(get('NOMBRE')).strip() if get('NOMBRE') else None
        if not nombre:
            continue
        # Algunos archivos traen el encabezado duplicado en la fila siguiente (p. ej. filas 2 y 3
        # ambas con 'ID'/'NOMBRE'/...); si no se filtra, esa fila se procesa como si fuera un
        # colaborador real llamado "NOMBRE", inflando conteos de retardo y colaboradores.
        if str(get('ID')).strip().upper() == 'ID' and nombre.upper() == 'NOMBRE':
            continue
        # Registros sin colaborador real identificado (placeholder tipo "NN-1"): no cuentan
        # como colaborador en ningún KPI ni listado.
        if es_nombre_placeholder(nombre):
            continue
        fecha = get('FECHA')
        entrada, salida = get('ENTRADA'), get('SALIDA')
        s_comer, r_comer = get('S. COMER'), get('R. COMER')
        entrada_ok = entrada not in (None, '0', 0)
        salida_ok = salida not in (None, '0', 0)
        s_comer_ok = s_comer not in (None, '0', 0)
        r_comer_ok = r_comer not in (None, '0', 0)

        es_sab = sab_activo and es_sabado(get('DIA'), fecha)
        if es_sab:
            # El HORARIO/HORARIO COMIDA que trae el archivo para sábado replica por error el de
            # entre semana (y varía por empleado); se ignora y se usa el horario real de grupo.
            horario_mostrado = sab_horario_label
            hrs = horas_trabajadas(entrada, salida, None if not sab_con_comida else s_comer,
                                    None if not sab_con_comida else r_comer) if (entrada_ok and salida_ok) else None
            mins_tarde = None
            if entrada_ok and sab_entrada_min is not None:
                e_min = to_minutes(entrada)
                if e_min is not None:
                    mins_tarde = e_min - sab_entrada_min
            r6, r13, r25 = clasifica_retardo_minutos(mins_tarde)
            fuera_horario = False
            if salida_ok and sab_salida_min is not None:
                s_min = to_minutes(salida)
                if s_min is not None and s_min > sab_salida_min + sab_tolerancia:
                    fuera_horario = True
        else:
            horario_mostrado = get('HORARIO')
            hrs = horas_trabajadas(entrada, salida, s_comer, r_comer) if (entrada_ok and salida_ok) else None
            r6 = bool(get('R 6-12')); r13 = bool(get('R 13-24')); r25 = bool(get('R +25'))
            fuera_horario = False

        # Marcaje del día completo: entre semana requiere las 4 marcaciones (entrada, salida y
        # comida); en sábado (sin comida oficial) basta con entrada y salida. Se usa para el KPI
        # de cumplimiento de horario/jornada, que respeta que sábado no tiene comida programada.
        registro_diario_completo = (entrada_ok and salida_ok) if es_sab else \
            (entrada_ok and salida_ok and s_comer_ok and r_comer_ok)
        # Falta real: el día no tiene ninguna checada (ni entrada ni salida) — se detecta aquí
        # directamente de las checadas diarias, sin depender de que la hoja "Días sin checadas"
        # lo haya reportado (algunos colaboradores no aparecen ahí aunque sí falten muchos días).
        falta_dia = (not entrada_ok) and (not salida_ok)
        # Checada incompleta: el colaborador sí se presentó ese día pero le faltó alguna marca
        # esperada. Entre semana, cualquiera de las 4 marcas ausente cuenta. En sábado no se
        # exige comida en un día normal (no hay comida programada), PERO si además ese sábado se
        # trabajó fuera del horario oficial (más de las 14:00), sí se espera que haya comida
        # checada — de lo contrario también cuenta como incompleta.
        if es_sab:
            checada_incompleta = (not falta_dia) and (
                (not (entrada_ok and salida_ok)) or (fuera_horario and not (s_comer_ok and r_comer_ok))
            )
        else:
            checada_incompleta = (not falta_dia) and (not registro_diario_completo)

        rec = {
            'unidad_negocio': meta['unidad_negocio'], 'marca': meta['marca'],
            'sucursal': meta['sucursal'], 'area': area_name,
            'id_empleado': emp_id, 'nombre': nombre,
            'fecha': str(fecha), 'dia_semana': get('DIA'),
            'horario_teorico': horario_mostrado,
            'entrada': entrada if entrada_ok else None,
            'salida': salida if salida_ok else None,
            'checada_completa': entrada_ok and salida_ok,
            'checada_incompleta': checada_incompleta,
            'falta_dia': falta_dia,
            'registro_diario_completo': registro_diario_completo,
            'horas_trabajadas': hrs,
            'jornada_excesiva_dia': bool(hrs and hrs > 12),
            'jornada_sabado': es_sab,
            'sabado_fuera_horario': fuera_horario,
            'retardo_6_12': r6,
            'retardo_13_24': r13,
            'retardo_mas_25': r25,
        }
        rec['retardo'] = rec['retardo_6_12'] or rec['retardo_13_24'] or rec['retardo_mas_25']
        records.append(rec)
    return records

def load_resumen_retardos(ws, meta):
    header_row, headers = find_header_row_generic(ws, ['ID', 'NOMBRE'])
    idx = {h: col_index(headers, h) for h in ['ID','NOMBRE','DEPARTAMENTO','TOTAL 6-12','TOTAL 13-24','TOTAL +25','SUMA']}
    out = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[idx['ID']] is None:
            continue
        if str(row[idx['ID']]).strip().upper() == 'ID':
            continue  # fila de encabezado duplicada, no un colaborador real
        if es_nombre_placeholder(row[idx['NOMBRE']]):
            continue  # registro sin colaborador real identificado (p. ej. "NN-1")
        out.append({
            'unidad_negocio': meta['unidad_negocio'],
            'id_empleado': normalize_emp_id(row[idx['ID']]),
            'nombre': row[idx['NOMBRE']],
            'departamento': row[idx['DEPARTAMENTO']] if idx.get('DEPARTAMENTO') is not None else None,
            'total_6_12': row[idx['TOTAL 6-12']] or 0,
            'total_13_24': row[idx['TOTAL 13-24']] or 0,
            'total_mas_25': row[idx['TOTAL +25']] or 0,
        })
    return out

def load_dias_sin_checadas(ws, meta):
    header_row, headers = find_header_row_generic(ws, ['ID', 'NOMBRE'])
    idx = {h: col_index(headers, h) for h in ['ID','NOMBRE','DEPARTAMENTO','TOTAL FALTAS','FECHAS']}
    out = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[idx['ID']] is None:
            continue
        if str(row[idx['ID']]).strip().upper() == 'ID':
            continue  # fila de encabezado duplicada, no un colaborador real
        if es_nombre_placeholder(row[idx['NOMBRE']]):
            continue  # registro sin colaborador real identificado (p. ej. "NN-1")
        fechas_raw = row[idx['FECHAS']] if idx.get('FECHAS') is not None else ''
        fechas = [f.strip() for f in str(fechas_raw).split(',')] if fechas_raw else []
        out.append({
            'unidad_negocio': meta['unidad_negocio'],
            'id_empleado': normalize_emp_id(row[idx['ID']]),
            'nombre': row[idx['NOMBRE']],
            'departamento': row[idx['DEPARTAMENTO']] if idx.get('DEPARTAMENTO') is not None else None,
            'total_faltas': row[idx['TOTAL FALTAS']] or 0,
            'fechas': fechas,
        })
    return out

def find_header_row_generic(ws, must_have):
    for r in range(1, 5):
        vals = [c.value for c in ws[r]]
        if vals and all(m in vals for m in must_have):
            return r, vals
    raise ValueError(f'No se encontró encabezado con {must_have}')

def iso_week(fecha_str):
    try:
        d = datetime.strptime(fecha_str, '%Y-%m-%d')
        y, w, _ = d.isocalendar()
        return f'{y}-W{w:02d}'
    except Exception:
        return None

def reconcile_faltas(daily, faltas_sheet, meta):
    """Cruza las faltas que trae la hoja 'Días sin checadas' con las faltas derivadas
    directamente de las checadas diarias (día sin entrada ni salida). Si un colaborador tiene
    días de falta real que esa hoja no reporta -o de plano no aparece ahí- quedan contabilizados
    de todas formas; es la unión de ambas fuentes, no solo la hoja de resumen."""
    info = {}
    for r in daily:
        if r.get('falta_dia'):
            e = info.setdefault(r['id_empleado'], {'nombre': r['nombre'], 'departamento': r.get('area'), 'fechas': set()})
            e['fechas'].add(r['fecha'])
    for f in faltas_sheet:
        e = info.setdefault(f['id_empleado'], {'nombre': f.get('nombre'), 'departamento': f.get('departamento'), 'fechas': set()})
        if not e.get('nombre'):
            e['nombre'] = f.get('nombre')
        if not e.get('departamento'):
            e['departamento'] = f.get('departamento')
        for fecha in (f.get('fechas') or []):
            if fecha:
                e['fechas'].add(fecha)
    out = []
    for emp_id, e in info.items():
        if not e['fechas']:
            continue
        out.append({
            'unidad_negocio': meta['unidad_negocio'], 'id_empleado': emp_id,
            'nombre': e['nombre'], 'departamento': e['departamento'],
            'total_faltas': len(e['fechas']), 'fechas': sorted(e['fechas']),
        })
    return out

def process_file(path, config):
    meta = parse_filename(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    daily, retardos_resumen, faltas_resumen = [], [], []
    meta['tiene_hoja_resumen_retardos'] = False
    meta['tiene_hoja_dias_sin_checadas'] = False
    for sheet in wb.sheetnames:
        s_upper = sheet.strip().upper()
        if s_upper == 'RESUMEN RETARDOS':
            meta['tiene_hoja_resumen_retardos'] = True
            retardos_resumen.extend(load_resumen_retardos(wb[sheet], meta))
        elif s_upper == 'DIAS SIN CHECADAS':
            meta['tiene_hoja_dias_sin_checadas'] = True
            faltas_resumen.extend(load_dias_sin_checadas(wb[sheet], meta))
        else:
            daily.extend(load_area_sheet(wb[sheet], sheet.strip(), meta, config))
    for r in daily:
        r['semana_iso'] = iso_week(r['fecha'])
    faltas_resumen = reconcile_faltas(daily, faltas_resumen, meta)
    return meta, daily, retardos_resumen, faltas_resumen

def aggregate_weekly(daily):
    from collections import defaultdict
    agg = defaultdict(lambda: {'horas': 0.0, 'dias_trabajados': 0})
    for r in daily:
        if r['horas_trabajadas'] is not None and r['semana_iso']:
            key = (r['unidad_negocio'], r['area'], r['id_empleado'], r['nombre'], r['semana_iso'])
            agg[key]['horas'] += r['horas_trabajadas']
            agg[key]['dias_trabajados'] += 1
    out = []
    for (unidad, area, emp_id, nombre, semana), v in agg.items():
        out.append({
            'unidad_negocio': unidad, 'area': area, 'id_empleado': emp_id,
            'nombre': nombre, 'semana_iso': semana,
            'horas_semana': round(v['horas'], 2), 'dias_trabajados': v['dias_trabajados']
        })
    return out

def apply_normativa(weekly, config, anio_ref):
    limite = config['jornada_ordinaria_maxima_semanal_por_anio'].get(str(anio_ref),
              config['jornada_ordinaria_maxima_semanal_por_anio']['2026'])
    tope_extra = config['horas_extra_maximas_semanales_por_anio'].get(str(anio_ref),
              config['horas_extra_maximas_semanales_por_anio']['2026'])
    for w in weekly:
        w['limite_semanal_aplicable'] = limite
        w['exceso_semanal'] = round(max(0, w['horas_semana'] - limite), 2)
        w['cumple_semana'] = w['horas_semana'] <= limite
        w['excede_tope_horas_extra'] = w['exceso_semanal'] > tope_extra
    return weekly

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--data-dir', required=True)
    ap.add_argument('--config', required=True)
    ap.add_argument('--out', required=True)
    ap.add_argument('--anio-ref', default='2026')
    args = ap.parse_args()

    with open(args.config, encoding='utf-8') as f:
        config = json.load(f)

    files = sorted(glob.glob(os.path.join(args.data_dir, '*.xlsx')))
    if not files:
        print('No se encontraron archivos .xlsx en', args.data_dir, file=sys.stderr)
        sys.exit(1)

    all_daily, all_retardos, all_faltas, unidades = [], [], [], []
    for path in files:
        try:
            meta, daily, retardos, faltas = process_file(path, config)
        except Exception as e:
            print(f'ERROR procesando {path}: {e}', file=sys.stderr)
            continue
        unidades.append(meta)
        all_daily.extend(daily)
        all_retardos.extend(retardos)
        all_faltas.extend(faltas)
        print(f'OK  {os.path.basename(path)} -> {meta["unidad_negocio"]}  ({len(daily)} registros diarios)')

    weekly = aggregate_weekly(all_daily)
    weekly = apply_normativa(weekly, config, args.anio_ref)

    dataset = {
        'generado': datetime.now().isoformat(),
        'anio_referencia': args.anio_ref,
        'normativa': config,
        'unidades_negocio': unidades,
        'registros_diarios': all_daily,
        'semanal': weekly,
        'retardos_resumen': all_retardos,
        'faltas_resumen': all_faltas,
    }
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(dataset, f, ensure_ascii=False)
    print(f'\nDataset consolidado escrito en {args.out}')
    print(f'Total registros diarios: {len(all_daily)} | Total semanas-empleado: {len(weekly)} '
          f'| Unidades: {len(unidades)}')

if __name__ == '__main__':
    main()
